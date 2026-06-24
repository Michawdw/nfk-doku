"""Erzeugt assets/templates.xlsx – einen Katalog mehrerer Bilddoku-Vorlagen.
Jeder Tab = eine Vorlage; der Tab-Name ist der Anzeigename in der App.
Spalten je Tab: Oberordner | Unterordner | Bildname | Pflichtanzahl
Neue Vorlage hinzufügen: einfach einen weiteren Tab anlegen und die Datei neu hochladen.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER = ["Oberordner", "Unterordner", "Bildname", "Pflichtanzahl"]

# Tab-Name -> Zeilen
TEMPLATES = {
    "NFK Vollverkabelung": [
        ("01_Anlieferung", "", "Materialanlieferung", 2),
        ("02_Serverraum", "NWS-Schrank", "Schrank_geschlossen", 1),
        ("02_Serverraum", "NWS-Schrank", "Schrank_offen", 1),
        ("02_Serverraum", "Patchpanel", "Patchfeld", 2),
        ("02_Serverraum", "", "Uebersicht_Serverraum", 1),
        ("03_Verkabelung", "Kassenzone", "Kabelweg_Kasse", 3),
        ("03_Verkabelung", "Lager", "Kabelweg_Lager", 3),
        ("03_Verkabelung", "", "Uebersicht_Verkabelung", 1),
        ("04_Accesspoints", "", "AP_Montage", 2),
        ("05_Abschluss", "", "Baustelle_aufgeraeumt", 1),
    ],
    "Nachkontrolle": [
        ("01_Maengel", "", "Maengel_vorher", 2),
        ("01_Maengel", "", "Maengel_behoben", 2),
        ("02_Abnahme", "", "Abnahme_Unterschrift", 1),
    ],
}

thin = Side(style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1F4E78")

wb = openpyxl.Workbook()
wb.remove(wb.active)  # Standard-Sheet entfernen

for tab, rows in TEMPLATES.items():
    ws = wb.create_sheet(title=tab)
    for c, title in enumerate(HEADER, start=1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            if c == 4:
                cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 14
    ws.freeze_panes = "A2"

out = "assets/templates.xlsx"
wb.save(out)
print("written:", out, "| Tabs:", list(TEMPLATES.keys()))
