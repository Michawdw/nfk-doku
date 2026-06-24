"""Erzeugt die mitgelieferte Beispiel-Struktur-Vorlage fuer die Bilddoku.
Spalten: Oberordner | Unterordner | Bildname | Pflichtanzahl
Einmalig ausgefuehrt; Ergebnis liegt als statisches Asset im Repo.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Struktur"

header = ["Oberordner", "Unterordner", "Bildname", "Pflichtanzahl"]
rows = [
    ("01_Anlieferung", "", "Materialanlieferung", 2),
    ("02_Serverraum", "NWS-Schrank", "Schrank_geschlossen", 1),
    ("02_Serverraum", "NWS-Schrank", "Schrank_offen", 1),
    ("02_Serverraum", "Patchpanel", "Patchfeld", 2),
    ("02_Serverraum", "", "Uebersicht_Serverraum", 1),
    ("03_Verkabelung", "Kassenzone", "Kabelweg_Kasse", 3),
    ("03_Verkabelung", "Lager", "Kabelweg_Lager", 3),
    ("03_Verkabelung", "", "Uebersicht_Verkabelung", 1),
    ("04_Accesspoints", "", "AP_Montage", 2),
    ("05_Abschluss", "", "Baustelle_aufgeraeumt", 1),
]

thin = Side(style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1F4E78")

for c, title in enumerate(header, start=1):
    cell = ws.cell(row=1, column=c, value=title)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

for r, row in enumerate(rows, start=2):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = border
        if c == 4:
            cell.alignment = Alignment(horizontal="center")

ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 26
ws.column_dimensions["D"].width = 14
ws.freeze_panes = "A2"

out = "assets/beispiel_bilddoku_template.xlsx"
wb.save(out)
print("written:", out)
