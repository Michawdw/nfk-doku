"""Verifiziert die ZIP-Patch-Logik aus js/bautagebuch.js gegen die echte Vorlage.
Repliziert setCell/patchSheet in Python, baut ein xlsx und prueft es mit openpyxl.
"""
import re, zipfile, shutil, io, datetime, os

SRC = "assets/vorlage_bautagebuch.xlsx"
OUT = "tools/_verify_out.xlsx"
SHEET = "xl/worksheets/sheet1.xml"

def excel_serial(datestr):
    y, mo, d = map(int, datestr.split("-"))
    return (datetime.date(y, mo, d) - datetime.date(1899, 12, 30)).days

def time_fraction(s):
    if not s: return None
    m = re.match(r"^(\d{1,2}):(\d{2})$", s.strip())
    if not m: return None
    h, mi = int(m.group(1)), int(m.group(2))
    return (h*60+mi)/1440

def esc(s):
    return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

def set_cell(xml, ref, build):
    pat = re.compile(r'<c r="%s"([^>]*?)(?:/>|>.*?</c>)' % re.escape(ref), re.S)
    if not pat.search(xml):
        print("  WARN Zelle nicht gefunden:", ref); return xml
    def repl(m):
        attrs = m.group(1)
        sm = re.search(r'\bs="(\d+)"', attrs)
        s = ' s="%s"' % sm.group(1) if sm else ''
        built = build()
        if built is None:
            return '<c r="%s"%s/>' % (ref, s)
        t = (' t="%s"' % built["t"]) if built.get("t") else ''
        return '<c r="%s"%s%s>%s</c>' % (ref, s, t, built["inner"])
    return pat.sub(repl, xml, count=1)

def text(v):
    def f():
        if v is None or v == "": return None
        return {"t":"inlineStr","inner":'<is><t xml:space="preserve">%s</t></is>' % esc(v)}
    return f

def num(v):
    def f():
        return None if v is None else {"inner":"<v>%s</v>" % v}
    return f

def patch(xml, model):
    xml = set_cell(xml, "G3", text(model["filiale"]))
    xml = set_cell(xml, "E5", num(excel_serial(model["datum"])))
    xml = set_cell(xml, "E6", text(model.get("beauftragung","NFK Vollverkabelung")))
    xml = set_cell(xml, "G7", num(model.get("anzTechniker")))
    for i in range(5):
        row = 8+i
        r = model["rows"][i] if i < len(model["rows"]) else {}
        xml = set_cell(xml, "B%d"%row, text(r.get("name")))
        xml = set_cell(xml, "I%d"%row, num(time_fraction(r.get("start"))))
        xml = set_cell(xml, "K%d"%row, num(time_fraction(r.get("ende"))))
        xml = set_cell(xml, "M%d"%row, num(time_fraction(r.get("pause"))))
        xml = set_cell(xml, "O%d"%row, text(r.get("bemerkung")))
    xml = set_cell(xml, "E14", text(model["taetigkeiten"]))
    xml = set_cell(xml, "E16", text(model["behinderungen"]))
    xml = set_cell(xml, "E19", text(model["vorkommnisse"]))
    xml = set_cell(xml, "B22", text(model["ortDatum"]))
    return xml

model = {
    "filiale":"7265 Memmingen", "datum":"2025-09-11",
    "beauftragung":"NFK Vollverkabelung", "anzTechniker":3,
    "rows":[
        {"name":"Lindner","start":"07:00","ende":"14:00","pause":"00:30","bemerkung":"x"},
        {"name":"Breternitz","start":"07:00","ende":"14:00","pause":"00:30"},
        {"name":"Martin","start":"07:00","ende":"14:00","pause":"00:30"},
    ],
    "taetigkeiten":"Zeile 1\nZeile 2 mit & < > Sonderzeichen\nZeile 3",
    "behinderungen":"Material fehlte",
    "vorkommnisse":"Keine",
    "ortDatum":"Memmingen 11.09.2025",
}

print("excel_serial(2025-09-11) =", excel_serial("2025-09-11"), "(erwartet 45911)")

# --- Patch anwenden: nur sheet1.xml ersetzen, Rest 1:1 kopieren ---
zin = zipfile.ZipFile(SRC, "r")
names = zin.namelist()
xml = zin.read(SHEET).decode("utf-8")
patched = patch(xml, model)

buf = io.BytesIO()
zout = zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED)
for n in names:
    data = patched.encode("utf-8") if n == SHEET else zin.read(n)
    zout.writestr(n, data)
zout.close(); zin.close()
with open(OUT,"wb") as f: f.write(buf.getvalue())

# --- Pruefen: Medien/Zeichnung erhalten? ---
zchk = zipfile.ZipFile(OUT,"r")
chk = zchk.namelist()
assert "xl/media/image1.png" in chk, "Logo verloren!"
assert "xl/drawings/drawing1.xml" in chk, "Zeichnung verloren!"
print("Logo + Zeichnung erhalten:", "xl/media/image1.png" in chk and "xl/drawings/drawing1.xml" in chk)
print("Dateien-Anzahl gleich:", len(names) == len(chk), "(%d/%d)"%(len(names),len(chk)))
zchk.close()

# --- Mit openpyxl lesen und Zelltypen/-werte pruefen ---
import openpyxl
wb = openpyxl.load_workbook(OUT)
ws = wb["Datum"]
def show(ref):
    c = ws[ref]
    print("  %-4s val=%r type=%s nfmt=%s" % (ref, c.value, c.data_type, c.number_format))
print("Zellpruefung:")
for ref in ["G3","E5","E6","G7","B8","I8","K8","M8","O8","B11","I11","E14","E16","E19","B22"]:
    show(ref)

# Validierungen
assert isinstance(ws["E5"].value, datetime.datetime) and ws["E5"].value.date() == datetime.date(2025,9,11), "Datum falsch"
assert ws["I8"].value == datetime.time(7,0), "Startzeit nicht als Uhrzeit"
assert ws["M8"].value == datetime.time(0,30), "Pause nicht als Dauer"
assert ws["G3"].value == "7265 Memmingen", "Filiale falsch"
assert ws["B8"].value == "Lindner", "Name falsch"
assert ws["B11"].value is None, "Zeile 11 sollte leer sein"
assert "Zeile 2 mit & < > Sonderzeichen" in ws["E14"].value, "Mehrzeiltext/Escaping falsch"
print("\nALLE PRUEFUNGEN OK")
print("Bild-Bytes identisch:", zipfile.ZipFile(SRC).read("xl/media/image1.png") == zipfile.ZipFile(OUT).read("xl/media/image1.png"))
